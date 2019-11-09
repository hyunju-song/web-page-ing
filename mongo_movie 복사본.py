from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)
client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.dbsparta

work_book = load_workbook('text.xlsx')
work_sheet = work_book['prac']

# 데이터를 입력합니다.
work_sheet.cell(row=1, column=1, value='순위')
work_sheet.cell(row=1, column=2, value='영화제목')
work_sheet.cell(row=1, column=3, value='평점')
work_sheet.cell(row=1, column=4, value='')


headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

movies = soup.select('table.list_ranking>tbody>tr')

row = 2
for movie in movies:
    ranking_image = movie.find('img')
    if ranking_image == None:
        continue

    ranking = ranking_image['alt']
    title = movie.select('.title> .tit5> a')[0].get_text()
    rating = movie.select('.point')[0].get_text()

    print(ranking, title, rating)

    db.movies.insert_one({'ranking':ranking, 'title':title, 'rating':rating})

    work_sheet.cell(row = row, column=1, value= ranking)
    work_sheet.cell(row = row, column=2, value= title)
    work_sheet.cell(row = row, column=3, value= rating)

    row +=1

all_users = list(db.movies.find())



